from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from tkinter.filedialog import Button, Label, asksaveasfilename, askopenfilename, BooleanVar, Checkbutton, Tk
import os


class Window:
    BIND_KEY = "<Button-1>"  # Button-1 is left click of mouse
    TEXT_LABEL = "-file doesn't choose-"
    BG = "white"  # back ground colour
    FONT = ('Ubuntu', 15)  # font properties
    PAGE_NAME = "Лист1"
    FG = "black"
    WIDTH_EL = 20
    RESOLUTION = "520x224"

    def __init__(self, master):
        master.title("Compare details")  # title window
        master.configure(bg=self.BG)  # back ground color of window
        master.geometry(self.RESOLUTION)
        self.containers_btn = [ContBtnAndLbl(master, 0, 0, "First file"),
                               ContBtnAndLbl(master, 1, 0, "Second file")]
        self.containers_check = [ContCheckBtn(master, 3, 0, "Cross"),
                                 ContCheckBtn(master, 4, 0, "Merge and remove re")]
        self.containers_empty = [ContEmptyLbl(master, 2, 0), ContEmptyLbl(master, 2, 1),
                                 ContEmptyLbl(master, 5, 0), ContEmptyLbl(master, 5, 1)]
        button = Button(master, text='Calculate and save', width=self.WIDTH_EL, font=self.FONT, bg=self.BG)
        button.grid(row=6, column=0, sticky="w")
        button.bind(self.BIND_KEY, self.check)  # bind button by func
        self.label = Label(master, width=self.WIDTH_EL+5, text="-status-", font=self.FONT, bg=self.BG, fg=self.FG)
        self.label.grid(row=6, column=1)
        self.lists = []
        self.res_merge_remove_re = []
        self.res_cross = []
        self.path_to_result_file = ""
        self.name_of_result_file = ""
        self.state_of_save = False

    def check(self, event):
        if all([c.state for c in self.containers_btn]) \
                and any([b.check_var.get() for b in self.containers_check]):
            self.save_file()
            if self.state_of_save:
                self.calculate()
        elif all([c.state for c in self.containers_btn]) \
                and not any([b.check_var.get() for b in self.containers_check]):
            self.error_message("Select type of calculation")
        else:
            self.error_message("Select files")

    def save_file(self):
        self.path_to_result_file = asksaveasfilename()
        try:
            if not self.path_to_result_file:
                self.state_of_save = False
                self.error_message("Empty path of result file")
            else:
                path_to_result_file = self.path_to_result_file.split('.')[0]
                name_of_result_file = path_to_result_file.split('/')[-1]
                self.name_of_result_file = name_of_result_file + str(".xlsx")
                self.path_to_result_file = path_to_result_file[:-len(name_of_result_file)]
                self.state_of_save = True

        except TypeError or FileNotFoundError:
            self.state_of_save = False
            self.error_message("File's not saved")

    def error_message(self, text_message):
        self.label['text'] = text_message
        self.label['fg'] = "red"

    def calculate(self):
        excel.create_file(self.path_to_result_file, self.name_of_result_file)
        path_to_first_file, path_to_second_file = [c.path for c in self.containers_btn]
        self.lists = [excel.read_file(path_to_first_file, self.PAGE_NAME),
                      excel.read_file(path_to_second_file, self.PAGE_NAME)]
        first_list, second_list = [c for c in self.lists]

        if self.containers_check[0].check_var.get() and not self.containers_check[1].check_var.get():
            self.res_cross = list(set(first_list) & set(second_list))
            excel.write_to_file(self.path_to_result_file, self.name_of_result_file,
                                self.res_cross)

        elif self.containers_check[1].check_var.get() and not self.containers_check[0].check_var.get():
            self.res_merge_remove_re = list(set(first_list + second_list))
            excel.write_to_file(self.path_to_result_file, self.name_of_result_file,
                                self.res_merge_remove_re)
        else:
            self.res_merge_remove_re = list(set(first_list + second_list))
            self.res_cross = list(set(first_list) & set(second_list))
            excel.write_to_file(self.path_to_result_file, self.name_of_result_file,
                                self.res_cross)
            excel.write_to_file(self.path_to_result_file, self.name_of_result_file,
                                self.res_merge_remove_re, 2)
        self.label['text'] = self.name_of_result_file + str(" saved!")
        self.label['fg'] = "green"


class HandlerExcel:
    def read_file(self, path, page_name):
        """read cells from file and page_name,
        then return all cells like one list
        """
        # open excel file
        wb = load_workbook(path)
        # choose page in excel file
        sheet = wb[page_name]

        # get max number of columns and rows
        column = str(get_column_letter(sheet.max_column))
        row = str(sheet.max_row)
        lst_of_el = []

        # read current of cells
        # cellObj -> it's cells from A1:column+row area
        for cellObj in sheet['A1':column + row]:
            # cell - cell from cellObj
            for cell in cellObj:
                # write cell in list
                lst_of_el.append(cell.value)
        return lst_of_el

    def create_file(self, path, file_name):
        wb = Workbook()
        wb.save(path+file_name)

    def write_to_file(self, path, file_name, writing_list, column=1):
        wb = load_workbook(path + file_name)
        sheet = wb["Sheet"]
        row = len(writing_list)
        for i in range(row):
            sheet.cell(i+1, column, value=writing_list[i])
        wb.save(filename=file_name)


class ContBtnAndLbl:
    ERROR_COLOR = "red"

    def __init__(self, master, row, column, text_button):
        self.state = False  # state of func
        button = Button(master, text=text_button, width=Window.WIDTH_EL, font=Window.FONT, bg=Window.BG)
        button.grid(row=row, column=column, sticky="w")
        button.bind(Window.BIND_KEY, self.button_event)  # bind button by func
        self.label = Label(master, width=Window.WIDTH_EL, text=Window.TEXT_LABEL,
                           font=Window.FONT, bg=Window.BG, fg=Window.FG)
        self.label.grid(row=row, column=column+1)
        self.path = ""

    def button_event(self, event):
        """Open window for get path on the file, filter .xls and .xlsx files, filter all errors"""
        self.path = askopenfilename()
        try:
            if not self.path:
                self.label['text'] = Window.TEXT_LABEL
                self.label['fg'] = self.ERROR_COLOR
                self.state = False
            else:
                if self.path.split('.')[-1].lower() in ['xls', 'xlsx']:
                    self.label['text'] = os.path.split(self.path)[-1]
                    self.label['fg'] = 'green'
                    # it'll return name of file into label
                    self.state = True
                else:
                    self.label['text'] = "-ErrTypeOfFile-"
                    self.label['fg'] = self.ERROR_COLOR
                    # if choose file which has different enhancing type, it'll return error into label
                    self.state = False
        except TypeError or FileNotFoundError:
            self.label['text'] = Window.TEXT_LABEL
            self.label['fg'] = self.ERROR_COLOR
            self.state = False
            # if return some errors, it'll be painted label in red colour


class ContCheckBtn:
    def __init__(self, master, row, column, text_btn):
        self.check_var = BooleanVar()
        check_button = Checkbutton(master, width=Window.WIDTH_EL-1, text=text_btn, font=Window.FONT,
                                   bg=Window.BG, variable=self.check_var)
        check_button.grid(row=row, column=column, sticky="n")


class ContEmptyLbl:
    def __init__(self, master, row, column):
        self.label = Label(master, width=Window.WIDTH_EL, text="---------------------",
                           font=Window.FONT, bg=Window.BG, fg=Window.FG)
        self.label.grid(row=row, column=column)


root = Tk()
excel = HandlerExcel()
gui = Window(root)
root.mainloop()
